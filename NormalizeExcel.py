from openpyxl import load_workbook
from pathlib import Path
import os

# Создаем директорию normalized, если её нет
os.makedirs('normalized', exist_ok=True)

# Фильтруем только .xlsx файлы
xls = [f for f in Path('Tables').iterdir() if f.is_file() and f.suffix.lower() == '.xlsx']

for xl in xls:
    try:
        print(f"Обработка файла: {xl.name}")
        wordbook = load_workbook(xl)
        
        sheet = wordbook.active
        
        merged = sheet.merged_cells.ranges
        hah = []
        for i in merged:
            hah.append(i)
        for i in hah:
            cellvalue = sheet[i.start_cell.coordinate].value
            sheet.unmerge_cells(range_string=i.coord)
            for row_index in range(i.min_row, i.max_row + 1):
                for col_index in range(i.min_col, i.max_col + 1):
                    sheet.cell(row=row_index, column=col_index, value=cellvalue)
        
        output_path = f'normalized/{xl.name}Course.xlsx'
        wordbook.save(output_path)
        print(f"Файл сохранен: {output_path}")
    except Exception as e:
        print(f"Ошибка при обработке файла {xl.name}: {e}")
        continue